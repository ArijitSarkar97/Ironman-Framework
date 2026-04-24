
import os
from openpyxl import load_workbook

def debug_excel_read(file_path):
    if not os.path.exists(file_path):
        print(f"File not found: {file_path}")
        return

    workbook = load_workbook(file_path)
    sheet = workbook.active
    
    print(f"Active Sheet: {sheet.title}")
    
    headers = [cell.value for cell in sheet[1]]
    print(f"Headers: {headers}")
    print(f"Header count: {len(headers)}")
    
    rows = list(sheet.iter_rows(min_row=2, values_only=True))
    print(f"Total rows (from min_row=2): {len(rows)}")
    
    for i, row in enumerate(rows):
        print(f"Row {i+2}: {row} (Length: {len(row)})")
        if any(row):
            print("  -> Has data")
        else:
            print("  -> EMPTY ROW")

if __name__ == "__main__":
    base_dir = r"c:\Users\Lenovo\PycharmProjects\Ironman-Framework"
    file_path = os.path.join(base_dir, "test-data", "checkout_test_data.xlsx")
    debug_excel_read(file_path)
