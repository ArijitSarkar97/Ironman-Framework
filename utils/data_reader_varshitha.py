from openpyxl import load_workbook
import os


def read_excel_data(file_path, sheet_name=None):
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"{file_path} not found")

    workbook = load_workbook(file_path)
    sheet = workbook[sheet_name] if sheet_name else workbook.active

    rows = list(sheet.iter_rows(values_only=True))
    headers = rows[0]

    data = []
    for row in rows[1:]:
        row_dict = dict(zip(headers, row))
        data.append(row_dict)

    return data